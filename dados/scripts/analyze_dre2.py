import openpyxl

file_path = 'dados/excel_exemplos/Dashboard_Financeiro_Exemplo.xlsx'
wb = openpyxl.load_workbook(file_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'=== ABA: {sheet_name} ===')
    print(f'Dimensões: {ws.dimensions}')
    print()
    # Mostrar primeiras 15 linhas
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=True), 1):
        print(f'Linha {i}: {row}')
    print()

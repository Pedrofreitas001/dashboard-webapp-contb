import openpyxl

file_path = 'dados/excel_exemplos/analise_despesas_exemplo.xlsx'
wb = openpyxl.load_workbook(file_path)

print('=== ABAS DO ARQUIVO ===')
print(f'Abas: {wb.sheetnames}')
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'=== ABA: {sheet_name} ===')
    print(f'Dimensões: {ws.dimensions}')
    print()
    # Mostrar primeiras 8 linhas e colunas
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8, values_only=True), 1):
        print(f'Linha {i}: {row}')
    print()
